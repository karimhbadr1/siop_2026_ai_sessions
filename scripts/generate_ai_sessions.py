import json
import re
from datetime import datetime, timedelta
from html import unescape
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Accepted Sessions with Submitter Info.xlsx"
OUTPUT = ROOT / "data" / "ai_sessions.json"
REVIEW_OUTPUT = ROOT / "data" / "symposium_review.csv"

KEYWORDS = re.compile(
    r"\b("
    r"ai|artificial intelligence|machine learning|ml|llm|large language model|"
    r"generative ai|genai|deep learning|neural network|natural language processing|nlp|"
    r"automation|automated|algorithmic|predictive analytics"
    r")\b",
    re.I,
)


def excel_serial_to_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    except Exception:
        return None


def format_label(dt):
    if not dt:
        return ""
    return f"{dt:%A} {dt.day} {dt:%B}, {dt:%I:%M %p}"


def format_time_only(dt):
    if not dt:
        return ""
    return dt.strftime("%H:%M")


def normalize(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def clean_markup(text):
    raw = normalize(text)
    if not raw:
        return ""
    raw = unescape(raw)
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = re.sub(r"</p>\s*<p>", "\n", raw, flags=re.I)
    raw = re.sub(r"<\s*br\s*/?\s*>", "\n", raw, flags=re.I)
    raw = re.sub(r"</?p[^>]*>", "\n", raw, flags=re.I)
    raw = re.sub(r"<[^>]+>", "", raw)
    raw = re.sub(r"[ \t]+", " ", raw)
    raw = re.sub(r"\n{3,}", "\n\n", raw)
    return raw.strip()


def split_symposium_citations(text):
    raw = clean_markup(text)
    if not raw:
        return ""

    marker = "[Symposium]."
    idx = raw.find(marker)
    if idx == -1:
        return raw

    head = raw[: idx + len(marker)].strip()
    tail = raw[idx + len(marker):].strip()
    if not tail:
        return head

    if "\n" in tail:
        parts = [p.strip() for p in tail.split("\n") if p.strip()]
        return "\n".join([head] + parts)

    # Insert line breaks only where a new paper citation clearly begins.
    tail = re.sub(r"\s+", " ", tail)
    split_pattern = re.compile(
        r"(?<=United States\.)\s*(?=[A-Z][A-Za-zÀ-ÿ'’\-\s,.&]*\(\d{4}\)\.)"
    )
    tail = split_pattern.sub("\n", tail)
    tail = re.sub(
        r"(?<=\(\d{4}\)\.)\s*(?=[A-Z][A-Za-zÀ-ÿ'’\-\s,.&]*\(\d{4}\)\.)",
        "\n",
        tail,
    )
    return "\n".join([head, tail.strip()])


def is_ai_session(row):
    text = " | ".join(
        [
            normalize(row.get("Session.Title")),
            normalize(row.get("SessionExtraData.Proposal Primary Content Area")),
            normalize(row.get("SessionExtraData.ProposalSecondaryContentArea")),
        ]
    )
    primary = normalize(row.get("SessionExtraData.Proposal Primary Content Area"))
    secondary = normalize(row.get("SessionExtraData.ProposalSecondaryContentArea"))
    return bool(KEYWORDS.search(text)) or primary == "Technology/Artificial Intelligence" or secondary == "Technology/Artificial Intelligence"


def clean_name(first, last):
    return " ".join(part for part in [normalize(first), normalize(last)] if part)


def main():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [normalize(cell.value) for cell in ws[1]]
    sessions = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if normalize(record.get("SessionAcceptStatus")) not in {"", "Accept"} and normalize(record.get("SessionExtraData.SessionAcceptStatus")) not in {"", "Accept"}:
            continue
        if not is_ai_session(record):
            continue

        start = excel_serial_to_datetime(record.get("Session.StartDateTime"))
        end = excel_serial_to_datetime(record.get("Session.EndDateTime"))
        citation_value = record.get("SessionExtraData.ProposalAPAStyleCitation")
        session_type = normalize(record.get("Session.Type"))

        sessions.append(
            {
                "session_id": normalize(record.get("Session.ID")),
                "title": normalize(record.get("Session.Title")),
                "type": session_type,
                "submitter": clean_name(record.get("Session.CreatorFirstName"), record.get("Session.CreatorLastName")),
                "citation": split_symposium_citations(citation_value) if session_type == "Symposium" else clean_markup(citation_value),
                "location": normalize(record.get("Session.Location")),
                "start_iso": start.isoformat() if start else "",
                "end_iso": end.isoformat() if end else "",
                "start_label": format_label(start),
                "end_label": format_label(end),
                "start_time_only": format_time_only(start),
                "end_time_only": format_time_only(end),
                "primary_area": normalize(record.get("SessionExtraData.Proposal Primary Content Area")),
                "secondary_area": normalize(record.get("SessionExtraData.ProposalSecondaryContentArea")),
            }
        )

    sessions.sort(key=lambda s: (s["start_iso"] or "", s["title"]))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"count": len(sessions), "sessions": sessions}, indent=2), encoding="utf-8")
    print(f"Wrote {len(sessions)} sessions to {OUTPUT}")

    symposium_rows = [s for s in sessions if s["type"] == "Symposium"]
    import csv

    with REVIEW_OUTPUT.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "session_id",
                "title",
                "start_label",
                "end_label",
                "citation",
                "location",
                "primary_area",
                "secondary_area",
            ],
        )
        writer.writeheader()
        for session in symposium_rows:
            writer.writerow(
                {
                    "session_id": session["session_id"],
                    "title": session["title"],
                    "start_label": session["start_label"],
                    "end_label": session["end_label"],
                    "citation": session["citation"],
                    "location": session["location"],
                    "primary_area": session["primary_area"],
                    "secondary_area": session["secondary_area"],
                }
            )

    print(f"Wrote {len(symposium_rows)} symposium rows to {REVIEW_OUTPUT}")


if __name__ == "__main__":
    main()
